"""
AMP-GEN Material Passport Extraction Task Starter
Extracts line items from a scanned BoQ PDF and fills the Material Passport template
"""

import os
import pandas as pd
import json
import re
from pathlib import Path

# pip install pytesseract pdf2image pillow openpyxl matplotlib pandas

try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError as e:
    print(f"Missing dependency: {e}. Install with: pip install pytesseract pdf2image pillow openpyxl matplotlib pandas")

# ============================================================================
# STEP 1: UNDERSTAND THE PDF STRUCTURE
# ============================================================================
"""
What to expect in the PDF:
- Scanned document (dot-matrix, hand-annotated)
- 64 line items, each representing building materials
- Each line typically contains:
  * Item number (1-64)
  * Description (material, size, quantity)
  * Unit (m, kg, sqm, etc.)
  * Quantity (numerical value)
  * Unit Rate (price per unit)
  * Amount (total cost)
  
The template columns are:
  - GREEN (REQUIRED): Item#, Description, Quantity, Unit, Floor, Discipline, MaterialCategory
  - AMBER (BONUS): Mass (kg), Carbon (kgCO2e), EPD_Source
  - GREY (out of scope): Cost-related fields
"""

# ============================================================================
# STEP 2: EXTRACT TEXT FROM PDF USING OCR
# ============================================================================

def extract_text_from_pdf(pdf_path, output_text_file='boq_extracted.txt'):
    """
    Convert PDF to images and extract text using OCR (pytesseract)
    This handles scanned PDFs.
    """
    print(f"Converting PDF to images: {pdf_path}")
    
    try:
        # Convert PDF pages to images
        images = convert_from_path(pdf_path, dpi=300)  # Higher DPI for better OCR
        
        all_text = []
        for i, image in enumerate(images):
            print(f"  Processing page {i+1}/{len(images)}...")
            
            # OCR using pytesseract
            text = pytesseract.image_to_string(image)
            all_text.append(f"\n--- PAGE {i+1} ---\n{text}")
        
        # Save raw OCR text for inspection
        with open(output_text_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(all_text))
        
        print(f"✓ OCR complete. Raw text saved to: {output_text_file}")
        return '\n'.join(all_text)
    
    except Exception as e:
        print(f"✗ OCR failed: {e}")
        print("  Troubleshooting:")
        print("  - Install Tesseract: brew install tesseract (macOS) or apt-get install tesseract-ocr (Linux)")
        print("  - Alternative: Use EasyOCR (pip install easyocr) or Claude's vision API")
        return None


def extract_text_easyocr(pdf_path, output_text_file='boq_extracted.txt'):
    """
    Alternative: Use EasyOCR if tesseract is not available
    pip install easyocr
    """
    try:
        import easyocr
        from pdf2image import convert_from_path
        
        print(f"Converting PDF to images: {pdf_path}")
        images = convert_from_path(pdf_path, dpi=300)
        
        # Initialize reader (downloads model on first use)
        reader = easyocr.Reader(['en'], gpu=False)
        
        all_text = []
        for i, image in enumerate(images):
            print(f"  Processing page {i+1}/{len(images)}...")
            
            # Save image temporarily for EasyOCR
            temp_path = f'/tmp/page_{i}.png'
            image.save(temp_path)
            
            # Extract text
            results = reader.readtext(temp_path)
            page_text = '\n'.join([text[1] for text in results])
            all_text.append(f"\n--- PAGE {i+1} ---\n{page_text}")
            
            os.remove(temp_path)
        
        full_text = '\n'.join(all_text)
        with open(output_text_file, 'w', encoding='utf-8') as f:
            f.write(full_text)
        
        print(f"✓ OCR complete. Raw text saved to: {output_text_file}")
        return full_text
    
    except ImportError:
        print("EasyOCR not installed. Use: pip install easyocr")
        return None


# ============================================================================
# STEP 3: PARSE LINE ITEMS FROM OCR TEXT
# ============================================================================

def parse_line_items(ocr_text):
    """
    Parse 64 line items from OCR'd text.
    
    This is CUSTOM to your specific PDF format.
    You'll need to inspect boq_extracted.txt and adjust the regex patterns.
    """
    
    line_items = []
    
    # Example regex patterns (ADJUST THESE based on your actual PDF format)
    # Pattern 1: "ItemNo | Description | Unit | Qty | Rate | Amount"
    # Pattern 2: Multiline format with line breaks
    
    lines = ocr_text.split('\n')
    
    current_item = {}
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Try to extract item number (usually starts with digits 1-64)
        match = re.match(r'^(\d{1,2})\s+(.+)', line)
        if match:
            if current_item:  # Save previous item
                line_items.append(current_item)
            
            item_num = match.group(1)
            description = match.group(2)
            
            current_item = {
                'Item#': int(item_num),
                'Description': description,
                'Quantity': None,
                'Unit': None,
                'Floor': 'TBD',
                'Discipline': 'TBD',
                'MaterialCategory': 'TBD',
            }
    
    if current_item:
        line_items.append(current_item)
    
    print(f"✓ Extracted {len(line_items)} line items")
    return line_items


# ============================================================================
# STEP 4: FILL THE EXCEL TEMPLATE
# ============================================================================

def fill_excel_template(template_path, line_items, output_path='output/passport_filled.xlsx'):
    """
    Load the template, add your extracted data, save the filled version
    """
    os.makedirs('output', exist_ok=True)
    
    print(f"Loading template: {template_path}")
    
    # Read template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Find where to start inserting data (skip examples, look for first empty row after headers)
    start_row = 5  # Adjust based on your template layout
    
    for idx, item in enumerate(line_items, start=start_row):
        ws[f'A{idx}'] = item.get('Item#', '')
        ws[f'B{idx}'] = item.get('Description', '')
        ws[f'C{idx}'] = item.get('Quantity', '')
        ws[f'D{idx}'] = item.get('Unit', '')
        ws[f'E{idx}'] = item.get('Floor', 'TBD')
        ws[f'F{idx}'] = item.get('Discipline', 'TBD')
        ws[f'G{idx}'] = item.get('MaterialCategory', 'TBD')
    
    wb.save(output_path)
    print(f"✓ Filled template saved to: {output_path}")
    
    return output_path


# ============================================================================
# STEP 5: EXPORT AS JSON
# ============================================================================

def export_to_json(line_items, output_path='output/passport.json'):
    """
    Export extracted data as JSON
    """
    os.makedirs('output', exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(line_items, f, indent=2, ensure_ascii=False)
    
    print(f"✓ JSON export saved to: {output_path}")


# ============================================================================
# STEP 6: CREATE VISUALIZATION
# ============================================================================

def create_visualization(line_items, output_path='output/visualization.html'):
    """
    Create an interactive visualization of material distribution
    """
    os.makedirs('output', exist_ok=True)
    
    try:
        import plotly.graph_objects as go
        import plotly.express as px
    except ImportError:
        print("Plotly not installed. Use: pip install plotly")
        return
    
    # Count materials by category
    df = pd.DataFrame(line_items)
    
    category_counts = df['MaterialCategory'].value_counts()
    
    # Create pie chart
    fig = px.pie(
        values=category_counts.values,
        names=category_counts.index,
        title='Material Distribution by Category (BoQ - CBRI Principal\'s Residence)',
        hole=0.3  # Donut chart
    )
    
    fig.write_html(output_path)
    print(f"✓ Visualization saved to: {output_path}")
    
    # Also create bar chart by discipline
    discipline_counts = df['Discipline'].value_counts()
    fig2 = px.bar(
        x=discipline_counts.index,
        y=discipline_counts.values,
        labels={'x': 'Discipline', 'y': 'Count'},
        title='Line Items by Discipline'
    )
    
    fig2.write_html(output_path.replace('.html', '_discipline.html'))
    print(f"✓ Discipline chart saved to: {output_path.replace('.html', '_discipline.html')}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    # Configure paths
    pdf_path = "BoQ_CBRI_Principals_Residence.pdf"  # Your downloaded file
    template_path = "AMP_Passport_Template.xlsx"     # Your template
    
    print("=" * 70)
    print("AMP-GEN Material Passport Extraction Task")
    print("=" * 70)
    
    # Step 1: Extract text from PDF
    if os.path.exists(pdf_path):
        print("\n[1/6] Extracting text from PDF using OCR...")
        ocr_text = extract_text_from_pdf(pdf_path)
        # Alternative if tesseract fails:
        # ocr_text = extract_text_easyocr(pdf_path)
        
        if not ocr_text:
            print("✗ OCR extraction failed. Please check manually and extract data.")
            exit(1)
    else:
        print(f"✗ PDF not found: {pdf_path}")
        print("  Download BoQ_CBRI_Principals_Residence.pdf from the email first")
        exit(1)
    
    # Step 2: Parse line items
    print("\n[2/6] Parsing line items...")
    line_items = parse_line_items(ocr_text)
    
    # IMPORTANT: Inspect boq_extracted.txt and manually adjust parse_line_items() function
    # to match your actual PDF format!
    
    # Step 3: Fill Excel template
    if os.path.exists(template_path):
        print("\n[3/6] Filling Excel template...")
        fill_excel_template(template_path, line_items)
    else:
        print(f"⚠ Template not found: {template_path}")
    
    # Step 4: Export as JSON
    print("\n[4/6] Exporting to JSON...")
    export_to_json(line_items)
    
    # Step 5: Create visualization
    print("\n[5/6] Creating visualization...")
    create_visualization(line_items)
    
    print("\n[6/6] Done!")
    print("\n" + "=" * 70)
    print("Next steps:")
    print("  1. Review boq_extracted.txt for OCR accuracy")
    print("  2. Adjust parse_line_items() regex patterns if needed")
    print("  3. Manually fill MaterialCategory, Discipline, Floor where OCR failed")
    print("  4. Commit code to GitHub (make multiple commits, not one huge one)")
    print("  5. Fill APPROACH.md explaining your process")
    print("  6. Update README.md with run instructions")
    print("=" * 70)
