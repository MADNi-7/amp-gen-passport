import os
import json
import time
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import google.generativeai as genai
from dotenv import load_dotenv
import openpyxl

# ============================================================================
# 1. DYNAMIC PATH RESOLUTION
# ============================================================================
SRC_DIR = Path(__file__).resolve().parent          
PROJECT_DIR = SRC_DIR.parent                      
WORKSPACE_DIR = PROJECT_DIR.parent                

load_dotenv(PROJECT_DIR / ".env")
load_dotenv(WORKSPACE_DIR / ".env")

OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def find_file(filename):
    search_paths = [
        WORKSPACE_DIR / "data" / filename,
        PROJECT_DIR / "data" / filename,
        PROJECT_DIR / filename,
        WORKSPACE_DIR / filename
    ]
    for path in search_paths:
        if path.exists():
            return path
    raise FileNotFoundError(f"Could not find '{filename}'.")

# ============================================================================
# 2. CONFIGURATION & KEYS
# ============================================================================
API_KEYS = [
    os.getenv("GEMINI_API_KEY"),
    os.getenv("GEMINI_API_KEY_2"),
    os.getenv("GEMINI_API_KEY_3")
]
API_KEYS = [k for k in API_KEYS if k]

if not API_KEYS:
    raise ValueError("No Gemini API keys found! Please add GEMINI_API_KEY to your .env file.")

MODELS = [
    "gemini-3.6-flash",
    "gemini-3.7-flash",
    "gemini-3.5-flash",
    "gemini-3.1-pro-preview"
]

EPD_DATABASE = {
    "Concrete": {"carbon_factor": 0.15, "source": "ICE Database V3.0"},
    "Steel": {"carbon_factor": 1.46, "source": "ICE Database V3.0"},
    "Brick": {"carbon_factor": 0.24, "source": "ICE Database V3.0"},
    "Timber": {"carbon_factor": 0.30, "source": "ICE Database V3.0"},
    "Glass": {"carbon_factor": 1.43, "source": "ICE Database V3.0"}
}

# ============================================================================
# 3. EXTRACTION FUNCTION 
# ============================================================================
def extract_data_from_pdf(pdf_path):
    prompt = """
    You are an expert Quantity Surveyor. Analyze this scanned Bill of Quantities (BoQ) and extract:
    1. 'metadata': The building block metadata from Page 1 (Project name, location, date, etc.)
    2. 'line_items': An array of exactly 64 line items. For each item, extract the following fields (use null if not explicitly found or inferable):
       - 'gmap_id': (String) GMAP Id (leave null)
       - 'id': (String) BOQ Item No.
       - 'article_number': (String) Article Number
       - 'external_db_id': (String) External DB Id
       - 'description': (String) Full description of the work/material
       - 'floor_section': (String) Floor / Section
       - 'discipline': (String) Infer if this is 'Civil', 'MEP', 'Architecture', etc.
       - 'material_product': (String) Material / Product
       - 'all_materials_detected': (String) All Materials Detected
       - 'material_category': (String) Infer the primary material (e.g., 'Concrete', 'Steel', 'Brick')
       - 'material_confidence': (String) Material Confidence
       - 'grade': (String) Grade
       - 'mix_ratio': (String) Mix Ratio (e.g., 1:4:8)
       - 'quantity': (Number) Original Quantity
       - 'unit': (String) Original Unit (e.g., Cum, Sqm, Kg)
       - 'volume_m3': (Number) Volume (m³)
       - 'area_m2': (Number) Area (m²)
       - 'length_m': (Number) Length (m)
       - 'weight_kg': (Number) Weight (kg)
       - 'count_nos': (Number) Count (Nos)
       - 'derived_quantity': (Number) Derived Quantity
       - 'derived_quantity_unit': (String) Derived Quantity Unit
       - 'derived_quantity_basis': (String) Derived Quantity Basis
       - 'schedule': (String) Schedule (DSR/SOR)
       - 'schedule_item_code': (String) Schedule Item Code
       - 'standard_code_reference': (String) Standard / Code Reference
       - 'classification': (String) Classification (Matched)
       - 'length_mm': (Number) Length (mm)
       - 'width_mm': (Number) Width (mm)
       - 'height_mm': (Number) Height (mm)
       - 'thickness_mm': (Number) Thickness (mm)
       - 'depth_mm': (Number) Depth (mm)
       - 'diameter_mm': (Number) Diameter (mm)
       - 'unit_rate': (Number) Unit Rate
       - 'total_cost': (Number) Total Cost
       - 'currency': (String) Currency
       - 'comment': (String) Comment
       
    Output STRICTLY in valid JSON format without markdown blocks.
    """
    key_idx, model_idx = 0, 0
    while True:
        curr_key, curr_model = API_KEYS[key_idx], MODELS[model_idx]
        genai.configure(api_key=curr_key)
        try:
            sample_file = genai.upload_file(path=str(pdf_path))
            model = genai.GenerativeModel(curr_model)
            response = model.generate_content([sample_file, prompt], request_options={"timeout": 600})
            raw_json = response.text.strip().removeprefix('```json').removesuffix('```')
            return json.loads(raw_json)
        except Exception as e:
            key_idx = (key_idx + 1) % len(API_KEYS)
            model_idx = (model_idx + 1) % len(MODELS)
            time.sleep(15)

# ============================================================================
# 4. PROCESS AND FILL TEMPLATE
# ============================================================================
def process_and_save(data, template_path):
    with open(OUTPUT_DIR / "building_meta.json", "w", encoding="utf-8") as f:
        json.dump(data.get("metadata", {}), f, indent=4)
        
    line_items = data.get("line_items", [])
    
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active
    
    # 🧹 ERASER: Completely wipe rows 4 through 10 to delete all template examples
    for r in range(4, 15):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None
            
    start_row = 4
    
    column_mapping = {
        'gmap_id': 1, 'id': 2, 'article_number': 3, 'external_db_id': 4,
        'description': 5, 'floor_section': 6, 'discipline': 7, 
        'material_product': 8, 'all_materials_detected': 9, 'material_category': 10,
        'material_confidence': 11, 'grade': 12, 'mix_ratio': 13, 
        'quantity': 14, 'unit': 15, 'volume_m3': 16, 'area_m2': 17,
        'length_m': 18, 'weight_kg': 19, 'count_nos': 20,
        'derived_quantity': 21, 'derived_quantity_unit': 22, 'derived_quantity_basis': 23,
        'schedule': 27, 'schedule_item_code': 28, 'standard_code_reference': 29, 
        'classification': 30, 'length_mm': 41, 'width_mm': 42, 
        'height_mm': 43, 'thickness_mm': 44, 'depth_mm': 45, 
        'diameter_mm': 46, 'unit_rate': 47, 'total_cost': 48, 
        'currency': 49, 'comment': 50
    }
    
    for idx, item in enumerate(line_items):
        current_row = start_row + idx
        for json_key, col_idx in column_mapping.items():
            val = item.get(json_key)
            if val is None or str(val).strip() == "" or val == "null":
                ws.cell(row=current_row, column=col_idx).value = "Not Specified"
            else:
                ws.cell(row=current_row, column=col_idx).value = val
                
        # AMBER Carbon calculation
        try:
            qty = float(item.get("quantity", 0))
        except (ValueError, TypeError):
            qty = 0
            
        category = item.get("material_category", "")
        if category in EPD_DATABASE:
            carbon_factor = EPD_DATABASE[category]["carbon_factor"]
            ws.cell(row=current_row, column=25).value = qty * carbon_factor
            ws.cell(row=current_row, column=26).value = carbon_factor
            ws.cell(row=current_row, column=50).value = f"EPD Source: {EPD_DATABASE[category]['source']}"
            
    excel_out = OUTPUT_DIR / "passport_filled.xlsx"
    wb.save(str(excel_out))
    
    with open(OUTPUT_DIR / "passport.json", "w", encoding="utf-8") as f:
        json.dump(line_items, f, indent=4)
        
    return pd.DataFrame(line_items)

# ============================================================================
# 5. VISUALIZATION
# ============================================================================
def create_visualization(df):
    if df.empty or "material_category" not in df.columns:
        return
    # Clean out any stray 'Not Specified' from the charts
    df_chart = df[df["material_category"] != "Not Specified"]
    
    plt.figure(figsize=(10, 6))
    mat_counts = df_chart["material_category"].value_counts()
    sns.barplot(x=mat_counts.values, y=mat_counts.index, palette="viridis")
    plt.title("Material Distribution across BoQ Line Items", fontsize=14)
    plt.xlabel("Number of Line Items", fontsize=12)
    plt.ylabel("Material Category", fontsize=12)
    plt.tight_layout()
    plt.savefig(str(OUTPUT_DIR / "visualization.png"))
    plt.close()

if __name__ == "__main__":
    pdf_file = find_file("BoQ_CBRI_Principals_Residence.pdf")
    template_file = find_file("AMP_Passport_Template.xlsx")
    extracted_data = extract_data_from_pdf(pdf_file)
    passport_df = process_and_save(extracted_data, template_file)
    create_visualization(passport_df)